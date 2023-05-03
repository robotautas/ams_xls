import os
import re
from openpyxl import load_workbook

file_pattern = re.compile(r"^results_\d{8}_\d{6}.xlsx$")


def get_excels():
    qty_dirs = 0
    filenames = []
    startswith_results_20 = 0
    startswith_2023 = 0
    print(os.listdir("./ams"))
    for dir in os.listdir("./ams"):
        if os.path.isdir("./ams/" + dir):
            qty_dirs += 1
            path = "./ams/" + dir
            for file in os.listdir(path):
                if file.endswith(".xlsx") or file.endswith(".xls"):
                    filenames.append(path + "/" + file)
                    if file.startswith("results_20"):
                        startswith_results_20 += 1
                    elif file.startswith("2023"):
                        startswith_2023 += 1
    filenames.sort()
    return filenames


def get_headers(file):
    wb = load_workbook(file, read_only=True)
    sheet = None
    if 'results(Age)' in wb.sheetnames:
        sheet = wb['results(Age)']
        for row in sheet.rows:
            if row[0].value == 'Sample_ID':
                headers = [cell.value for cell in row if cell.value and cell.value != 'Age(R1)']
                return tuple(headers)





def get_structure_patterns(files):
    patterns = {}
    for file in files:
        try:
            headers = get_headers(file)
            print(headers)
            if headers in patterns:
                patterns[headers] += 1
            else: 
                patterns[headers] = 1
        except:
            print(f"could not process {file}")


    print(patterns)

get_structure_patterns(get_excels())
