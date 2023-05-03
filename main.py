from openpyxl import load_workbook
import os
import re

STANDARTS = []
"""
one directory represents one measurement results 
this class will operate in a single directory
"""


class DirManager:
    file_pattern = re.compile(r"^results_\d{8}_\d{6}.xlsx$")

    def __init__(self, path):
        self.path = path

    def get_excel_file(self):
        """for now, most common systemic data format is expected"""
        files = os.listdir(self.path)
        datafile = None
        for file in files:
            if re.match(self.file_pattern, file):
                datafile = file
                break
        wb = load_workbook(self.path + datafile, read_only=True)
        sheet = wb.get_sheet_by_name('')
        if not wb.
        return datafile
    
    def chk_file_integrity(self):
        file = ms


# wb = load_workbook(filename="./ams/20201015/results_20201018_165811.xlsx", read_only=True)
# print(wb.get_sheet_names())

dm = DirManager("./ams/20201015/")
print(dm.check_for_excel_file())
